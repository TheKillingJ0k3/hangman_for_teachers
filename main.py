#! python3

from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, Border, Side
from random import randint
import os, shutil
from PIL import ImageTk, Image


#GLOBAL VARIABLES
word_selected = ''
hidden_word = ''
tkinter_letter_selected = ''
tkinter_hidden_word = ''
number_of_letters = ''
attempts = ''
icon = 1
tkinter_attempts = ''
english_alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
german_alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'ä', 'ö', 'ü', 'ß']
greek_alphabet = ['α', 'β', 'γ', 'δ', 'ε', 'ζ', 'η', 'θ', 'ι', 'κ', 'λ', 'μ', 'ν', 'ξ', 'ο', 'π', 'ρ', 'σ', 'τ', 'υ', 'φ', 'χ', 'ψ', 'ω']
levels_to_columns = {'A1':'1',
                    'A2':'2',
                    'B1':'3',
                    'B2':'4',
                    'C1':'5',
                    'C2':'6'
                    }

#####################################################
# TODO: one tab for each language

##################################  FUNCTIONS  ##################################################

#creates  folder inside wd, if it doesn't already exist
def createFolder(path):
    try:
        if not os.path.exists(path):
            os.mkdir(path)
    except OSError:
        print('Error creating directory' + path)

################## styles ##################
def excel_styles(excel_sheet):
    Arial_11_Font = Font(name='Arial', size=11)

    Arial_11_bold_Font = Font(name='Arial', size=11, bold=True)
    for columnNum in range(1, excel_sheet.max_column + 1):
        excel_sheet.cell(row=1, column=columnNum).font = Arial_11_bold_Font
    excel_sheet.freeze_panes = 'A2'
############################################

def open_excel(excel_file_path): # creates excel the first time the game runs
    try:
        wb = openpyxl.load_workbook(excel_file_path, data_only=True) # , data_only=True in case file has a lot of formulas
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'A1' # or cell.value = skata
        ws.column_dimensions['A'].width = 25
        ws['B1'] =  'A2'
        ws.column_dimensions['B'].width = 25
        ws['C1'] =  'B1'
        ws.column_dimensions['C'].width = 25
        ws['D1'] =  'B2'
        ws.column_dimensions['D'].width = 25
        ws['E1'] =  'C1'
        ws.column_dimensions['E'].width = 25
        ws['F1'] =  'C2'
        ws.column_dimensions['F'].width = 25
        excel_styles(ws)
    wb.save(excel_file_path)
    return wb, ws

######################################################
def show_words():
    os.startfile(r".\\Hangman Excel\\Hangman Excel.xlsx") # , data_only=True in case file has a lot of formulas


def start_game(level):
    open_game_frame()
    # Label.destroy()
    global word_selected
    global hidden_word
    global tkinter_hidden_word
    global tkinter_attempts
    word_selected = None
    while word_selected == None:
        word_selected = ws.cell(row=(randint(2, ws.max_row)), column=int(levels_to_columns[level])).value
    # print(word_selected)
    hidden_word = word_selected.split()[1]
    hidden_word = word_selected.split()[0] + ' ' + word_selected.split()[1][0] + ' ' + (len(hidden_word)-2)*'_ ' + word_selected.split()[1][-1]
    print(hidden_word)
    tkinter_hidden_word.set(hidden_word)
    Label(word_frame, textvariable=tkinter_hidden_word).pack(side = LEFT, anchor=W)
    
    if word_selected.split()[1][-2] in german_alphabet:
        for letter in german_alphabet:
            if german_alphabet.index(letter) < 6:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 1,column=german_alphabet.index(letter), sticky=W)
            elif german_alphabet.index(letter) >= 6 and german_alphabet.index(letter) < 12:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 2,column=(german_alphabet.index(letter)-6), sticky=W)
            elif german_alphabet.index(letter) >= 12 and german_alphabet.index(letter) < 18:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 3,column=(german_alphabet.index(letter)-12), sticky=W)
            elif german_alphabet.index(letter) >= 18  and german_alphabet.index(letter) < 24:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 4,column=(german_alphabet.index(letter)-18), sticky=W)
            elif german_alphabet.index(letter) >= 24:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 5,column=(german_alphabet.index(letter)-24), sticky=W)
    elif word_selected.split()[1][-2] in greek_alphabet:
        for letter in greek_alphabet:
            if greek_alphabet.index(letter) < 6:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 1,column=greek_alphabet.index(letter), sticky=W)
            elif greek_alphabet.index(letter) >= 6 and greek_alphabet.index(letter) < 12:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 2,column=(greek_alphabet.index(letter)-6), sticky=W)
            elif greek_alphabet.index(letter) >= 12 and greek_alphabet.index(letter) < 18:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 3,column=(greek_alphabet.index(letter)-12), sticky=W)
            elif greek_alphabet.index(letter) >= 18  and greek_alphabet.index(letter) < 24:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 4,column=(greek_alphabet.index(letter)-18), sticky=W)
            elif greek_alphabet.index(letter) >= 24:
                ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter)).grid(row= 5,column=(greek_alphabet.index(letter)-24), sticky=W)
    Label(attempt_frame, textvariable=tkinter_attempts).pack(side = LEFT, anchor=SE)
    return word_selected, hidden_word, tkinter_hidden_word

######################################################
def start_level_A1():
    global attempts
    global tkinter_attempts
    for ele in word_frame.winfo_children():
        ele.pack_forget()
    for ele in attempt_frame.winfo_children():
        ele.pack_forget()
    level_selected = 'A1'
    attempts = 5
    tkinter_attempts.set('attempts: ' + (str(attempts)))
    start_game(level_selected)

def start_level_A2():
    global attempts
    global tkinter_attempts
    for ele in word_frame.winfo_children():
        ele.pack_forget()
    for ele in attempt_frame.winfo_children():
        ele.pack_forget()
    level_selected = 'A2'
    attempts = 5
    tkinter_attempts.set('attempts: ' + (str(attempts)))
    start_game(level_selected)

def start_level_B1():
    global attempts
    global tkinter_attempts
    for ele in word_frame.winfo_children():
        ele.pack_forget()
    for ele in attempt_frame.winfo_children():
        ele.pack_forget()
    level_selected = 'B1'
    attempts = 10
    tkinter_attempts.set('attempts: ' + (str(attempts)))
    start_game(level_selected)

def start_level_B2():
    global attempts
    global tkinter_attempts
    for ele in word_frame.winfo_children():
        ele.pack_forget()
    for ele in attempt_frame.winfo_children():
        ele.pack_forget()
    level_selected = 'B2'
    attempts = 10
    tkinter_attempts.set('attempts: ' + (str(attempts)))
    start_game(level_selected)

def start_level_C1():
    global attempts
    global tkinter_attempts
    for ele in word_frame.winfo_children():
        ele.pack_forget()
    for ele in attempt_frame.winfo_children():
        ele.pack_forget()
    level_selected = 'C1'
    attempts = 15
    tkinter_attempts.set('attempts: ' + (str(attempts)))
    start_game(level_selected)

def start_level_C2():
    global attempts
    global tkinter_attempts
    for ele in word_frame.winfo_children():
        ele.pack_forget()
    for ele in attempt_frame.winfo_children():
        ele.pack_forget()
    level_selected = 'C2'
    attempts = 15
    tkinter_attempts.set('attempts: ' + (str(attempts)))
    start_game(level_selected)
######################################################

def select_letter(letter): # play round
    global attempts
    global icon
    global tkinter_attempts
    global tkinter_letter_selected
    global word_selected
    global hidden_word
    global tkinter_hidden_word
    tkinter_letter_selected = letter
    if tkinter_letter_selected in word_selected:
        print("Nice!")
        article = (word_selected.split())[0]
        word_selected = list(word_selected.split()[1]) # no article
        hidden_word = (hidden_word.split())[1:]
        # print(word_selected)
        # print(hidden_word)
        for i in range(1,len(word_selected)-1): #1, -1 because we don't care about first and last letter
            if tkinter_letter_selected == word_selected[i]:
                hidden_word[i] = tkinter_letter_selected
        # print(hidden_word)
        tkinter_hidden_word.set(article + ' ' + ' '.join(hidden_word))
        word_selected = article + ' ' + ''.join(word_selected)
        hidden_word = article + ' ' + ' '.join(hidden_word)
        tkinter_letter_selected = ''
        # print(word_selected)
        # print(hidden_word)
        #disable button
        ttk.Button(letter_frame, text=str(letter), width=2, command = lambda letter=letter: select_letter(letter))['state'] = DISABLED # not working
    else:
        print("Try again!")
        attempts = attempts -1
        print(attempts)
        tkinter_attempts.set('attempts: ' + str(attempts))
        icon = icon + 1
        if icon < 9 :
            original_image = Image.open(".\\hangman{}.png".format(icon))
            original_image = original_image.resize((350,300), Image.Resampling.LANCZOS)
            original_image = ImageTk.PhotoImage(original_image)
            img = Label(hangman_frame, image = original_image)
            img.image = original_image
            img.place(x=0, y=0)
        if attempts == 0 :
            messagebox.showinfo(title='Fail!', message='GAME OVER')
    # print(word_selected.split())
    # print(hidden_word.split())
    if ''.join(word_selected.split()) == ''.join(hidden_word.split()) and attempts > 0 :
        print('Congratulations!')
        messagebox.showinfo(title='Success!', message='Congratulations!')
    elif ''.join(word_selected.split()) != ''.join(hidden_word.split()) and attempts > 0 :
        print('One more!')
    else:
        messagebox.showinfo(title='Dude, stop!', message='How about playing a new game instead?') # .geometry(f"dimension+{root.winfo_x()}+{root.winfo_y()}")
        # root_x = root.winfo_rootx()
        # root_y = root.winfo_rooty()
        # win_x = root_x + 300
        # win_y = root_y + 100
        # win.geometry(f'+{win_x}+{win_y}')
    # return tkinter_letter_selected, tkinter_hidden_word, word_selected, hidden_word


def function(): # blank function for buttons not used
    pass

###################### GUI Functions ##############################

def open_game_frame():
    game_frame.pack_forget()
    hangman_frame.pack_forget()
    word_frame.pack_forget()
    letter_attempts_frame.pack_forget()
    letter_frame.pack_forget()
    attempt_frame.pack_forget()
    game_frame.pack(side=LEFT, fill='both', expand=1) #  fill='both', expand=1
    hangman_frame.pack(side=TOP)
    
    original_image = Image.open(".\\hangman1.png")
    original_image = original_image.resize((350,300), Image.Resampling.LANCZOS)
    original_image = ImageTk.PhotoImage(original_image)
    img = Label(hangman_frame, image = original_image)
    img.image = original_image
    img.place(x=0, y=0)
    
    word_frame.pack(side=BOTTOM)
    letter_attempts_frame.pack(side=RIGHT)
    letter_frame.pack(side=TOP)
    attempt_frame.pack(side=BOTTOM)
    frame.pack_forget()

#############################################################################################

createFolder('.\\Hangman Excel')
wb, ws = open_excel('.\\Hangman Excel\Hangman Excel.xlsx')

###################### main GUI - Button creation #########################################
root = Tk()
root.title('Hangman')
root.geometry('500x350')
# root.state('zoomed')
# root.option_add('*tear0ff', False) #opens fullscreen
# root.iconbitmap('.\\hangman.ico')

tkinter_hidden_word = StringVar()
tkinter_letter_selected = StringVar()
tkinter_attempts = StringVar()

# background_image = PhotoImage(file='C:\\Users\\kj\\Documents\\Python Projects\\Comic downloader\\crowd-img.png')
# background_label = Label(root, image=background_image)
# background_label.place(x=0, y=0, relwidth=1, relheight=1)

frame = Frame(root, borderwidth=5, relief="sunken", width=500, height=500) # 100 -200
frame.pack()

game_frame = Frame(root, borderwidth=3, relief="sunken", width=350, height=350, bg='pink') # 100 -200
hangman_frame = Frame(game_frame, borderwidth=3, relief="sunken", width=350, height=300) # 100 -200
word_frame = Frame(game_frame, borderwidth=3, relief="sunken", width=350, height=150) # 100 -200
letter_attempts_frame = Frame(root, borderwidth=2, relief="sunken", width=150, height=350, bg='yellow') # 100 -200
letter_frame = Frame(letter_attempts_frame, borderwidth=2, relief="sunken", width=150, height=320, bg='yellow') # 100 -200
attempt_frame = Frame(letter_attempts_frame, borderwidth=2, relief="sunken", width=150, height=30, bg='red') # 100 -200

##############################   MENU  #############################################
menubar = Menu(root) #creates menubar
root.config(menu = menubar) #same as frame['menu'] = menubar, doesn't need menu=menu_file etc inside cascade

#creating submenus in frame menu/menubar
data_menu = Menu(menubar, tearoff=False) #first_lineises new submenu
# menubar.add_cascade(label='View Data', menu=data_menu) #creates name of new submenu
# data_menu.add_command(label='Buses', command=open_Buses_window) # if command function has argument, window opens automatically


action_menu = Menu(menubar, tearoff=False) #first_lineises new submenu
menubar.add_cascade(label='New Game', menu=action_menu) #creates name of new submenu
level_menu = Menu(action_menu, tearoff=False)
level_menu.add_command(label='A1', command=start_level_A1) #adds option to submenu
level_menu.add_command(label='A2', command=start_level_A2) #adds option to submenu
level_menu.add_command(label='B1', command=start_level_B1) #adds option to submenu
level_menu.add_command(label='B2', command=start_level_B2) #adds option to submenu
level_menu.add_command(label='C1', command=start_level_C1) #adds option to submenu
level_menu.add_command(label='C2', command=start_level_C2) #adds option to submenu
action_menu.add_cascade(label='Choose Level', menu=level_menu) #creates name of new submenu

settings_menu = Menu(menubar, tearoff=False)
menubar.add_cascade(label='Settings', menu=settings_menu) #creates name of new submenu
settings_menu.add_command(label='Show Saved Words', command=show_words)
settings_menu.add_command(label='Add Words', command=show_words) #creates name of new submenu
# word_settings_menu = Menu(settings_menu, tearoff=False)
# word_settings_menu.add_command(label='A1', command=show_words) #adds option to submenu

advice_menu = Menu(menubar, tearoff=False) #first_lineises new submenu
menubar.add_cascade(label='Advice', menu=advice_menu) #creates name of new submenu
advice_menu.add_command(label='1', command=function) #adds option to submenu
# menu_app.add_command(Label='Save app', command=save_excel) #adds option to submenu

menu_exit = Menu(menubar)
menubar.add_cascade(label='Exit', command=frame.quit)
######################################################################################

root.mainloop()